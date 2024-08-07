## build-in
import os
import sys

## 3rd Pty.
# import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# from openpyxl.cell.cell import Cell

# [DEF] Environment variables
__dir_name__ = os.path.basename(os.path.dirname(__file__))
WORKSPACE_HOME = os.path.dirname(__file__).replace(f"/experiments/{__dir_name__}", "")
""".../ipirec"""

# [SET] Env.append($WORKSPACE_HOME)
sys.path.append(WORKSPACE_HOME)

## custom modules
from core import DirectoryPathValidator


class NMFResultAggregator:
    def __init__(self) -> None:
        self._model_name_ = "nmf"
        self.params_dict = {
            "DataSet": set(),
            "Dtype": set(),
            "factor_dim": set(),
            "factor_iter": set(),
            "learning_rate": set(),
            "generalization": set(),
            "train_iter": set(),
            "FrobNorm": set(),
        }
        """
        - Summary: Conditions_dictionary
        - Elements:
            - Key: 
                - DataSet
                - Dtype
                - factor_dim
                - factor_iter
                - learning_rate
                - generalization
                - train_iter
                - FrobNorm
            - Value: params_set (set)
        """
        self.parsed_results_list = list()
        self.properties_key_list = list()
        """k, Top-N, Metrics = [, ... ]"""
        # self.properties_key_list.append("")
        self.properties_key_list.append("k")
        # self.properties_key_list.append("Top-N")

    # end : init()

    def aggregation(
        self,
        results_dir_path: str,
        save_workbook_path: str,
        filtering: bool = True,
    ):
        ## 경로확인
        # RESULTS_PATH
        results_dir_path = (
            os.path.dirname(results_dir_path)
            if os.path.isfile(results_dir_path)
            else results_dir_path
        )
        self._target_dir_path = results_dir_path
        if not DirectoryPathValidator.exist_dir(results_dir_path):
            raise FileNotFoundError()

        # WORKBOOK_PATH
        workbook_dir_path = os.path.dirname(save_workbook_path)
        if not DirectoryPathValidator.exist_dir(workbook_dir_path):
            DirectoryPathValidator.mkdir(workbook_dir_path)

        # 파일확인, 결과 불러오기
        ext_str = ".csv"
        for file_path in [
            f"{results_dir_path}/{fp}"
            for fp in os.listdir(results_dir_path)
            if ext_str in fp
        ]:
            lines: list = None
            with open(
                file=file_path,
                mode="rt",
                encoding="utf-8",
            ) as fin:
                lines = fin.readlines()
                fin.close()
            # end : StreamReader()

            # 결과 집계 및 변환
            results_lines = self.__parse_result__(
                lines=lines,
                filtering=filtering,
            )
            # append elements;
            self.parsed_results_list.extend(results_lines)
        # end : for (results)

        # 결과출력
        self.__dump_results__(file_path=save_workbook_path)

    # end : public Any aggregation()

    def __dump_results__(
        self,
        file_path: str,
    ) -> None:
        workbook = Workbook()
        __dir_path = (
            self._target_dir_path if file_path == "" else os.path.dirname(file_path)
        )
        __file_name = (
            f"parsed_{self._model_name_}_results.xlsx"
            if file_path == ""
            else os.path.basename(file_path)
        )
        if not DirectoryPathValidator.exist_dir(__dir_path):
            DirectoryPathValidator.mkdir(__dir_path)

        ## Conditions sheet
        # Key, values = [value1, value2, ... ] \n
        conditions_sheet: Worksheet = workbook.create_sheet(
            title=f"{self._model_name_}_params",
            index=0,
        )
        # [header]
        conditions_sheet.append(["Properties", "Descriptions"])
        # [values]
        for key, values in self.params_dict.items():
            properties_list = list()
            properties_list.append(key)
            """
            for value in values:
                if isinstance(value, float):
                    properties_list.append(value)
                elif isinstance(value, int):
                    properties_list.append(value)
                else:
                    value: str
                    properties_list.append(value)
            """
            values = sorted(values)
            properties_list.extend(values)
            conditions_sheet.append(properties_list)
        # end : for (model_params)

        ## Results sheet
        results_sheet: Worksheet = workbook.create_sheet(
            title="raw_results",
            index=1,
        )
        # [header]
        results_sheet.append(self.properties_key_list)
        # [values]
        for result_list in self.parsed_results_list:
            results_sheet.append(result_list)
        # end : for (results)

        # export
        file_path = f"{__dir_path}/{__file_name}"
        workbook.save(filename=file_path)

    # end : private void dump_results()

    def __parse_result__(
        self,
        lines: list,
        filtering: bool,
    ) -> list:
        parsed_list = list()

        # DATA
        split_ch = ","
        dataset = lines[1].split(split_ch)[1].strip()
        decision = lines[1].split(split_ch)[2].strip()
        set_id = int(lines[2].split(split_ch)[1])
        self.params_dict["DataSet"].add(dataset)
        self.params_dict["Dtype"].add(decision)

        # FACTORIZER
        split_ch = ":"
        factor_dim = int(lines[4].split(split_ch)[1].strip())
        factor_iter = int(lines[5].split(split_ch)[1].strip())
        self.params_dict["factor_dim"].add(factor_dim)
        self.params_dict["factor_iter"].add(factor_iter)

        # TRAIN
        learning_rate = float(lines[7].split(split_ch)[1].strip())
        generalization = float(lines[8].split(split_ch)[1].strip())
        train_iter = int(lines[9].split(split_ch)[1].strip())
        frob_norm = int(lines[10].split(split_ch)[1].strip())
        self.params_dict["learning_rate"].add(learning_rate)
        self.params_dict["generalization"].add(generalization)
        self.params_dict["train_iter"].add(train_iter)
        self.params_dict["FrobNorm"].add(frob_norm)

        # RESULTS
        split_ch = ","

        # [HEADER] Sheet
        if len(self.properties_key_list) < 2:
            metrics_properties = [
                l.strip() for l in lines[11].split(split_ch) if l != ""
            ]
            self.properties_key_list.extend(metrics_properties)
            self.properties_key_list[1] = "Top-N"
            ## Fixed Prop. Seq.
            self.properties_key_list.append("DataSet")
            self.properties_key_list.append("Dtype")
            self.properties_key_list.append("factor_dim")
            self.properties_key_list.append("factor_iter")
            self.properties_key_list.append("learning_rate")
            self.properties_key_list.append("generalization")
            self.properties_key_list.append("train_iter")
            self.properties_key_list.append("FrobNorm")
        # end : if

        # keys = [k.strip() for k in lines[11].split(split_ch)]
        # keys[0] = "idx"
        for idx in range(12, len(lines)):
            # values = [v.strip() for v in lines[idx].split(split_ch)]
            values = [float(v) for v in lines[idx].split(split_ch)]
            values.pop(0)

            result_list = list()
            result_list.append(set_id)
            result_list.extend(values)
            if int(values[5]) == 0 and filtering:
                continue
            # end : if (results_filtering)

            result_list.append(dataset)
            result_list.append(decision)
            result_list.append(factor_dim)
            result_list.append(factor_iter)
            result_list.append(learning_rate)
            result_list.append(generalization)
            result_list.append(train_iter)
            result_list.append(frob_norm)
            parsed_list.append(result_list)
            """
            line = f"{set_id},"
            for value in values:
                line += f"{value},"
            # end : for (metrics)
            line += str.format(
                "{0},{1},{2},{3},{4},{5},{6},{7}\n",
                dataset,
                decision,
                factor_dim,
                factor_iter,
                learning_rate,
                generalization,
                train_iter,
                frob_norm,
            )
            parsed_list.append(line)
            """
        # end : for (top_n_items_metrics)

        return parsed_list

    # end : private list parse_reuslt()


# end : class


if __name__ == "__main__":
    ## [NMF] Aggr
    # RESULTS_DIR_PATH = "/Users/taegyu.hwang/Desktop/results/20240510/NMF"
    # workbook_file_path = "/Users/taegyu.hwang/Desktop/results/20240510/nmf_summary.xlsx"

    ## [NMF] AggrRev
    RESULTS_DIR_PATH = "/Users/taegyu.hwang/Documents/tghwang_git_repo/ipirec/results/20240614/1556/overall/NMF"
    workbook_file_path = "/Users/taegyu.hwang/Documents/tghwang_git_repo/ipirec/results/20240614/1556/overall/nmf_results.xlsx"

    inst = NMFResultAggregator()
    inst.aggregation(
        results_dir_path=RESULTS_DIR_PATH,
        save_workbook_path=workbook_file_path,
    )

# end : main()
